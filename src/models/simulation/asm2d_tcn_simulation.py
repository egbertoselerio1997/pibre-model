"""ASM2d-TCN workbook, matrix, and reduced steady-state simulation helpers."""

from __future__ import annotations

import os
from concurrent.futures import ProcessPoolExecutor
from pathlib import Path
from typing import Any, Mapping

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from src.utils.simulation import (
    get_repo_root,
    load_model_params,
    load_paths_config,
    render_simulation_artifact_paths,
    save_simulation_artifacts,
)


MODEL_NAME = "asm2d_tcn_simulation"
WORKBOOK_PATH_KEY = "asm2d_tcn_reference_workbook"
DATA_PATTERN_KEY = "asm2d_tcn_simulation_data_pattern"
METADATA_PATTERN_KEY = "asm2d_tcn_simulation_metadata_pattern"
STOICHIOMETRIC_SHEET_NAME = "stoichiometric_matrix"
COMPOSITION_SHEET_NAME = "composition_matrix"
PARAMETER_SHEET_NAME = "parameter_table"
PARAMETER_VALUE_COLUMN_INDEX = 5

HEADER_FILL = PatternFill(fill_type="solid", fgColor="D8DEE6")
SECTION_FILL = PatternFill(fill_type="solid", fgColor="EEF2F5")
HEADER_FONT = Font(bold=True, color="22303C")

STOICHIOMETRIC_COEFFICIENTS: list[dict[str, dict[str, str]]] = [
    {"coefficients": {"S_F": "1-{f_SI}", "S_I": "{f_SI}", "X_S": "-1"}},
    {"coefficients": {"S_F": "1-{f_SI}", "S_I": "{f_SI}", "X_S": "-1"}},
    {"coefficients": {"S_F": "1-{f_SI}", "S_I": "{f_SI}", "X_S": "-1"}},
    {"coefficients": {"S_F": "1-{f_SI}", "S_I": "{f_SI}", "X_S": "-1"}},
    {"coefficients": {"S_F": "-1/{Y_H}", "S_O2": "1-1/{Y_H}", "X_H": "1"}},
    {"coefficients": {"S_A": "-1/{Y_H}", "S_O2": "1-1/{Y_H}", "X_H": "1"}},
    {
        "coefficients": {
            "S_F": "-1/{Y_H}",
            "S_NO2": "(1-{Y_H})/((8/7)*{Y_H})",
            "S_NO3": "-((1-{Y_H})/((8/7)*{Y_H}))",
            "X_H": "1",
        }
    },
    {
        "coefficients": {
            "S_F": "-1/{Y_H}",
            "S_N2": "(1-{Y_H})/(1.72*{Y_H})",
            "S_NO2": "-((1-{Y_H})/(1.72*{Y_H}))",
            "X_H": "1",
        }
    },
    {
        "coefficients": {
            "S_A": "-1/{Y_H}",
            "S_NO2": "(1-{Y_H})/((8/7)*{Y_H})",
            "S_NO3": "-((1-{Y_H})/((8/7)*{Y_H}))",
            "X_H": "1",
        }
    },
    {
        "coefficients": {
            "S_A": "-1/{Y_H}",
            "S_N2": "(1-{Y_H})/(1.72*{Y_H})",
            "S_NO2": "-((1-{Y_H})/(1.72*{Y_H}))",
            "X_H": "1",
        }
    },
    {"coefficients": {"S_A": "1", "S_F": "-1"}},
    {"coefficients": {"X_I": "{f_XIBM}", "X_S": "1-{f_XIBM}", "X_H": "-1"}},
    {"coefficients": {"S_A": "-1", "X_PAO": "-{Y_PAO}", "X_PHA": "1"}},
    {"coefficients": {"S_O2": "-{Y_PHA}", "X_PP": "1", "X_PHA": "-{Y_PHA}"}},
    {
        "coefficients": {
            "S_NO2": "{Y_PHA}/(8/7)",
            "S_NO3": "-({Y_PHA}/(8/7))",
            "X_PP": "1",
            "X_PHA": "-{Y_PHA}",
        }
    },
    {
        "coefficients": {
            "S_N2": "{Y_PHA}/1.72",
            "S_NO2": "-({Y_PHA}/1.72)",
            "X_PP": "1",
            "X_PHA": "-{Y_PHA}",
        }
    },
    {"coefficients": {"S_O2": "1-1/{Y_PAO}", "X_PAO": "1", "X_PHA": "-1/{Y_PAO}"}},
    {
        "coefficients": {
            "S_NO2": "(1-{Y_PAO})/((8/7)*{Y_PAO})",
            "S_NO3": "-((1-{Y_PAO})/((8/7)*{Y_PAO}))",
            "X_PAO": "1",
            "X_PHA": "-1/{Y_PAO}",
        }
    },
    {
        "coefficients": {
            "S_N2": "(1-{Y_PAO})/(1.72*{Y_PAO})",
            "S_NO2": "-((1-{Y_PAO})/(1.72*{Y_PAO}))",
            "X_PAO": "1",
            "X_PHA": "-1/{Y_PAO}",
        }
    },
    {"coefficients": {"X_I": "{f_XIBM}", "X_S": "1-{f_XIBM}", "X_PAO": "-1"}},
    {"coefficients": {"X_PP": "-1"}},
    {"coefficients": {"S_A": "1", "X_PHA": "-1"}},
    {"coefficients": {"S_NO2": "1/{Y_AOB}", "S_O2": "-((3.43-{Y_AOB})/{Y_AOB})", "X_AOB": "1"}},
    {"coefficients": {"S_NO2": "-1/{Y_NOB}", "S_NO3": "1/{Y_NOB}", "S_O2": "-((1.14-{Y_NOB})/{Y_NOB})", "X_NOB": "1"}},
    {"coefficients": {"X_I": "{f_XIBM}", "X_S": "1-{f_XIBM}", "X_AOB": "-1"}},
    {"coefficients": {"X_I": "{f_XIBM}", "X_S": "1-{f_XIBM}", "X_NOB": "-1"}},
    {"coefficients": {"S_PO4": "-1", "X_TSS": "1.42", "X_MeOH": "-3.45", "X_MeP": "4.87"}},
    {"coefficients": {"S_PO4": "1", "X_TSS": "-1.42", "X_MeOH": "3.45", "X_MeP": "-4.87"}},
]

COMPOSITION_FORMULAS: dict[str, dict[str, str]] = {
    "S_A": {"COD": "1"},
    "S_F": {"COD": "1", "TN": "{i_NSF}", "TKN": "{i_NSF}", "TP": "{i_PSF}"},
    "S_I": {"COD": "1", "TN": "{i_NSI}", "TKN": "{i_NSI}", "TP": "{i_PSI}"},
    "S_NH4": {"TN": "1", "TKN": "1"},
    "S_NO2": {"TN": "1"},
    "S_NO3": {"TN": "1"},
    "S_PO4": {"TP": "1"},
    "X_I": {"COD": "1", "TN": "{i_NXI}", "TKN": "{i_NXI}", "TP": "{i_PXI}", "VSS": "{i_VSS_XI}"},
    "X_S": {"COD": "1", "TN": "{i_NXS}", "TKN": "{i_NXS}", "TP": "{i_PXS}", "VSS": "{i_VSS_XS}"},
    "X_H": {"COD": "1", "TN": "{i_NBM}", "TKN": "{i_NBM}", "TP": "{i_PBM}", "VSS": "{i_VSS_BM}"},
    "X_PAO": {"COD": "1", "TN": "{i_NBM}", "TKN": "{i_NBM}", "TP": "{i_PBM}", "VSS": "{i_VSS_BM}"},
    "X_PP": {"TP": "1"},
    "X_PHA": {"COD": "1", "VSS": "{i_VSS_PHA}"},
    "X_AOB": {"COD": "1", "TN": "{i_NBM}", "TKN": "{i_NBM}", "TP": "{i_PBM}", "VSS": "{i_VSS_BM}"},
    "X_NOB": {"COD": "1", "TN": "{i_NBM}", "TKN": "{i_NBM}", "TP": "{i_PBM}", "VSS": "{i_VSS_BM}"},
    "X_TSS": {"TSS": "1"},
    "X_MeP": {"TP": "{i_PMeP}"},
}

NITROGEN_CONTINUITY_TERMS = {
    "S_F": "{i_NSF}",
    "S_I": "{i_NSI}",
    "S_N2": "1",
    "S_NO2": "1",
    "S_NO3": "1",
    "X_I": "{i_NXI}",
    "X_S": "{i_NXS}",
    "X_H": "{i_NBM}",
    "X_PAO": "{i_NBM}",
    "X_AOB": "{i_NBM}",
    "X_NOB": "{i_NBM}",
}

PHOSPHORUS_CONTINUITY_TERMS = {
    "S_F": "{i_PSF}",
    "S_I": "{i_PSI}",
    "X_I": "{i_PXI}",
    "X_S": "{i_PXS}",
    "X_H": "{i_PBM}",
    "X_PAO": "{i_PBM}",
    "X_PP": "1",
    "X_AOB": "{i_PBM}",
    "X_NOB": "{i_PBM}",
    "X_MeP": "{i_PMeP}",
}

TSS_CONTINUITY_TERMS = {
    "X_I": "{i_TSS_i}",
    "X_S": "{i_TSS_i}",
    "X_H": "{i_TSS_i}",
    "X_PAO": "{i_TSS_i}",
    "X_PP": "{i_TSS_PP}",
    "X_PHA": "{i_TSS_PHA}",
    "X_AOB": "{i_TSS_i}",
    "X_NOB": "{i_TSS_i}",
    "X_MeOH": "1",
    "X_MeP": "1",
}


def load_asm2d_tcn_simulation_params(repo_root: str | Path | None = None) -> dict[str, Any]:
    """Load the configured ASM2d-TCN simulation definition."""

    return load_model_params(MODEL_NAME, repo_root)


def resolve_asm2d_tcn_workbook_path(
    repo_root: str | Path | None = None,
    *,
    paths_config: Mapping[str, Any] | None = None,
) -> Path:
    """Resolve the configured canonical workbook path."""

    root = get_repo_root(repo_root)
    config = dict(paths_config) if paths_config is not None else load_paths_config(root)
    return root / Path(config[WORKBOOK_PATH_KEY])


def resolve_asm2d_tcn_simulation_artifact_paths(
    repo_root: str | Path | None = None,
    *,
    timestamp: str | None = None,
    paths_config: Mapping[str, Any] | None = None,
) -> tuple[Path, Path, str]:
    """Resolve the configured ASM2d-TCN dataset and metadata output paths."""

    return render_simulation_artifact_paths(
        MODEL_NAME,
        repo_root=repo_root,
        timestamp=timestamp,
        paths_config=paths_config,
        data_pattern_key=DATA_PATTERN_KEY,
        metadata_pattern_key=METADATA_PATTERN_KEY,
    )


def create_asm2d_tcn_workbook(
    workbook_path: str | Path | None = None,
    *,
    repo_root: str | Path | None = None,
    model_params: Mapping[str, Any] | None = None,
) -> Path:
    """Create the canonical ASM2d-TCN workbook with formula-driven matrices."""

    workbook = build_asm2d_tcn_workbook(model_params=model_params, repo_root=repo_root)
    output_path = Path(workbook_path) if workbook_path is not None else resolve_asm2d_tcn_workbook_path(repo_root)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path.resolve()


def build_asm2d_tcn_workbook(
    *,
    model_params: Mapping[str, Any] | None = None,
    repo_root: str | Path | None = None,
) -> Workbook:
    """Build the workbook object for the configured ASM2d-TCN reference model."""

    params = dict(model_params) if model_params is not None else load_asm2d_tcn_simulation_params(repo_root)
    workbook_config = _validate_workbook_config(params)
    parameter_refs = _build_parameter_reference_map(workbook_config["parameters"])

    workbook = Workbook()
    stoichiometric_sheet = workbook.active
    stoichiometric_sheet.title = STOICHIOMETRIC_SHEET_NAME
    composition_sheet = workbook.create_sheet(COMPOSITION_SHEET_NAME)
    parameter_sheet = workbook.create_sheet(PARAMETER_SHEET_NAME)

    _write_stoichiometric_sheet(stoichiometric_sheet, workbook_config, parameter_refs)
    _write_composition_sheet(composition_sheet, workbook_config, parameter_refs)
    _write_parameter_sheet(parameter_sheet, workbook_config["parameters"])

    for worksheet in workbook.worksheets:
        _auto_size_columns(worksheet)
        worksheet.auto_filter.ref = worksheet.dimensions

    return workbook


def get_asm2d_tcn_matrices(model_params: Mapping[str, Any] | None = None) -> dict[str, Any]:
    """Build numeric Petersen and composition matrices for the configured ASM2d-TCN model."""

    params = dict(model_params) if model_params is not None else load_asm2d_tcn_simulation_params()
    runtime = _validate_runtime_structure(params)
    workbook_config = runtime["workbook_config"]
    parameter_values = _build_parameter_value_map(workbook_config["parameters"])
    state_columns = list(runtime["state_columns"])
    measured_output_columns = list(runtime["measured_output_columns"])
    process_names = list(runtime["process_names"])
    process_types = list(runtime["process_types"])
    state_index = _build_state_index(state_columns)
    output_index = _build_state_index(measured_output_columns)

    petersen_matrix = np.zeros((len(process_names), len(state_columns)), dtype=float)
    composition_matrix = np.zeros((len(measured_output_columns), len(state_columns)), dtype=float)

    for row_index, process_definition in enumerate(STOICHIOMETRIC_COEFFICIENTS):
        row_values = petersen_matrix[row_index]
        direct_coefficients = process_definition["coefficients"]

        for state_name, expression in direct_coefficients.items():
            row_values[state_index[state_name]] = _evaluate_numeric_expression(expression, parameter_values)

        row_values[state_index["S_NH4"]] = -sum(
            row_values[state_index[state_name]] * _evaluate_numeric_expression(factor_expression, parameter_values)
            for state_name, factor_expression in NITROGEN_CONTINUITY_TERMS.items()
        )

        if "S_PO4" not in direct_coefficients:
            row_values[state_index["S_PO4"]] = -sum(
                row_values[state_index[state_name]] * _evaluate_numeric_expression(factor_expression, parameter_values)
                for state_name, factor_expression in PHOSPHORUS_CONTINUITY_TERMS.items()
            )

        row_values[state_index["S_ALK"]] = (
            row_values[state_index["S_NH4"]] / 14.0
            - row_values[state_index["S_NO2"]] / 14.0
            - row_values[state_index["S_NO3"]] / 14.0
            + row_values[state_index["S_PO4"]] / 31.0
        )

        if "X_TSS" not in direct_coefficients:
            row_values[state_index["X_TSS"]] = sum(
                row_values[state_index[state_name]] * _evaluate_numeric_expression(factor_expression, parameter_values)
                for state_name, factor_expression in TSS_CONTINUITY_TERMS.items()
            )

    for output_name in measured_output_columns:
        output_row = output_index[output_name]
        for state_name, mapping in COMPOSITION_FORMULAS.items():
            if output_name in mapping:
                composition_matrix[output_row, state_index[state_name]] = _evaluate_numeric_expression(
                    mapping[output_name],
                    parameter_values,
                )

    return {
        "petersen_matrix": petersen_matrix,
        "composition_matrix": composition_matrix,
        "process_names": process_names,
        "process_types": process_types,
        "state_index": state_index,
        "state_columns": state_columns,
        "measured_output_columns": measured_output_columns,
    }


def build_asm2d_tcn_metadata(
    model_params: Mapping[str, Any],
    *,
    sample_count: int,
    random_seed: int,
    dataset_file: str | None = None,
) -> dict[str, Any]:
    """Create the metadata contract for the ASM2d-TCN composite-only dataset."""

    runtime = _validate_runtime_structure(model_params)
    state_columns = list(runtime["state_columns"])
    measured_output_columns = list(runtime["measured_output_columns"])
    process_names = list(runtime["process_names"])
    process_types = list(runtime["process_types"])
    operational_columns = list(runtime["operational_columns"])
    dependent_columns = [f"Out_{name}" for name in measured_output_columns]

    return {
        "simulation_name": MODEL_NAME,
        "n_samples": sample_count,
        "random_seed": random_seed,
        "dependent_columns": dependent_columns,
        "independent_columns": operational_columns + [f"In_{name}" for name in state_columns],
        "identifier_columns": [],
        "ignored_columns": [],
        "dataset_file": dataset_file,
        "state_columns": state_columns,
        "measured_output_columns": measured_output_columns,
        "operational_columns": operational_columns,
        "processes": process_names,
        "process_types": process_types,
        "petersen_matrix_shape": [len(process_names), len(state_columns)],
        "composition_matrix_shape": [len(measured_output_columns), len(state_columns)],
        "schema_version": str(model_params["schema_version"]),
    }


def generate_asm2d_tcn_dataset(
    *,
    model_params: Mapping[str, Any] | None = None,
    n_samples: int | None = None,
    random_seed: int | None = None,
    parallel_workers: int | None = None,
    parallel_chunk_size: int | None = None,
) -> tuple[pd.DataFrame, dict[str, Any], dict[str, Any]]:
    """Generate a reduced steady-state ASM2d-TCN dataset with composite-only outputs."""

    params = dict(model_params) if model_params is not None else load_asm2d_tcn_simulation_params()
    runtime = _validate_runtime_structure(params)
    configured_hyperparameters = params["hyperparameters"]
    sample_count = int(n_samples if n_samples is not None else configured_hyperparameters["n_samples"])
    if sample_count < 0:
        raise ValueError("n_samples must be greater than or equal to 0.")

    seed = int(random_seed if random_seed is not None else configured_hyperparameters["seed"])
    requested_parallel_workers = int(
        parallel_workers if parallel_workers is not None else configured_hyperparameters.get("parallel_workers", 1)
    )
    requested_parallel_chunk_size = int(
        parallel_chunk_size
        if parallel_chunk_size is not None
        else configured_hyperparameters.get("parallel_chunk_size", sample_count or 1)
    )

    matrix_bundle = get_asm2d_tcn_matrices(params)
    state_columns = list(runtime["state_columns"])
    operational_columns = list(runtime["operational_columns"])
    measured_output_columns = list(runtime["measured_output_columns"])

    influent_states = np.zeros((sample_count, len(state_columns)), dtype=float)
    operational = np.zeros((sample_count, len(operational_columns)), dtype=float)
    measured_outputs = np.zeros((sample_count, len(measured_output_columns)), dtype=float)
    chunk_results: list[tuple[int, np.ndarray, np.ndarray, np.ndarray]] = []
    if sample_count > 0:
        worker_count = _resolve_parallel_workers(requested_parallel_workers, sample_count)
        chunk_size = min(_resolve_parallel_chunk_size(requested_parallel_chunk_size), sample_count)
        chunk_specs = [
            {
                "chunk_start": chunk_start,
                "chunk_size": min(chunk_size, sample_count - chunk_start),
                "chunk_seed": seed + chunk_index,
                "model_params": params,
                "matrix_bundle": matrix_bundle,
                "runtime": runtime,
            }
            for chunk_index, chunk_start in enumerate(range(0, sample_count, chunk_size))
        ]

        if worker_count == 1 or len(chunk_specs) == 1:
            chunk_results = [_generate_asm2d_tcn_dataset_chunk(**chunk_spec) for chunk_spec in chunk_specs]
        else:
            with ProcessPoolExecutor(max_workers=worker_count) as executor:
                futures = [executor.submit(_generate_asm2d_tcn_dataset_chunk, **chunk_spec) for chunk_spec in chunk_specs]
                chunk_results = [future.result() for future in futures]

    for chunk_start, chunk_influent, chunk_operational, chunk_measured in chunk_results:
        chunk_end = chunk_start + len(chunk_influent)
        influent_states[chunk_start:chunk_end] = chunk_influent
        operational[chunk_start:chunk_end] = chunk_operational
        measured_outputs[chunk_start:chunk_end] = chunk_measured

    influent_df = pd.DataFrame(influent_states, columns=[f"In_{name}" for name in state_columns])
    operational_df = pd.DataFrame(operational, columns=operational_columns)
    measured_df = pd.DataFrame(measured_outputs, columns=[f"Out_{name}" for name in measured_output_columns])
    dataset = pd.concat([operational_df, influent_df, measured_df], axis=1)

    metadata = build_asm2d_tcn_metadata(
        params,
        sample_count=sample_count,
        random_seed=seed,
    )

    return dataset, metadata, matrix_bundle


def run_asm2d_tcn_simulation(
    *,
    save_artifacts: bool = True,
    repo_root: str | Path | None = None,
    n_samples: int | None = None,
    random_seed: int | None = None,
    parallel_workers: int | None = None,
    parallel_chunk_size: int | None = None,
    timestamp: str | None = None,
) -> dict[str, Any]:
    """Run the ASM2d-TCN reduced steady-state simulation and optionally persist artifacts."""

    params = load_asm2d_tcn_simulation_params(repo_root)
    dataset, metadata, matrix_bundle = generate_asm2d_tcn_dataset(
        model_params=params,
        n_samples=n_samples,
        random_seed=random_seed,
        parallel_workers=parallel_workers,
        parallel_chunk_size=parallel_chunk_size,
    )

    artifact_paths: dict[str, Path | None] = {
        "dataset_csv": None,
        "metadata_json": None,
    }

    if save_artifacts:
        dataset_path, metadata_path, persisted_metadata = save_simulation_artifacts(
            dataset,
            metadata,
            MODEL_NAME,
            repo_root=repo_root,
            timestamp=timestamp,
            data_pattern_key=DATA_PATTERN_KEY,
            metadata_pattern_key=METADATA_PATTERN_KEY,
        )
        metadata = persisted_metadata
        artifact_paths = {
            "dataset_csv": dataset_path,
            "metadata_json": metadata_path,
        }

    return {
        "dataset": dataset,
        "metadata": metadata,
        "petersen_matrix": matrix_bundle["petersen_matrix"],
        "composition_matrix": matrix_bundle["composition_matrix"],
        "matrix_bundle": matrix_bundle,
        "composite_matrix": matrix_bundle["composition_matrix"],
        "artifact_paths": artifact_paths,
    }


def _validate_workbook_config(model_params: Mapping[str, Any]) -> dict[str, Any]:
    if "workbook" not in model_params:
        raise KeyError("asm2d_tcn_simulation must define a workbook section.")

    workbook_config = dict(model_params["workbook"])
    expected_sheets = [STOICHIOMETRIC_SHEET_NAME, COMPOSITION_SHEET_NAME, PARAMETER_SHEET_NAME]
    configured_sheets = list(workbook_config["sheets"])
    if configured_sheets != expected_sheets:
        raise ValueError("asm2d_tcn_simulation workbook sheets must match the required three-sheet contract.")

    dissolved_state_columns = list(workbook_config["dissolved_state_columns"])
    particulate_state_columns = list(workbook_config["particulate_state_columns"])
    state_columns = list(workbook_config["state_columns"])
    composite_variables = list(workbook_config["composite_variables"])
    processes = list(workbook_config["processes"])
    parameter_rows = list(workbook_config["parameters"])
    state_units = dict(workbook_config["state_units"])

    _validate_unique_names(dissolved_state_columns, "dissolved_state_columns")
    _validate_unique_names(particulate_state_columns, "particulate_state_columns")
    _validate_unique_names(state_columns, "state_columns")
    _validate_unique_names(composite_variables, "composite_variables")

    if state_columns != dissolved_state_columns + particulate_state_columns:
        raise ValueError("asm2d_tcn_simulation state_columns must concatenate dissolved and particulate state columns.")

    missing_state_units = [state_name for state_name in state_columns if state_name not in state_units]
    if missing_state_units:
        missing_display = ", ".join(missing_state_units)
        raise KeyError(f"asm2d_tcn_simulation missing state units for: {missing_display}")

    if len(processes) != len(STOICHIOMETRIC_COEFFICIENTS):
        raise ValueError("asm2d_tcn_simulation workbook process count does not match the stoichiometric matrix definition.")

    process_indices = [int(process["index"]) for process in processes]
    if process_indices != list(range(1, len(processes) + 1)):
        raise ValueError("asm2d_tcn_simulation workbook processes must be sequentially indexed from 1.")

    parameter_names = [str(parameter_row["excel_name"]) for parameter_row in parameter_rows]
    _validate_unique_names(parameter_names, "parameter excel_name")

    for parameter_row in parameter_rows:
        for required_key in ("category", "symbol", "excel_name", "description", "value", "unit"):
            if required_key not in parameter_row:
                raise KeyError(f"asm2d_tcn_simulation parameter row missing '{required_key}'.")
        float(parameter_row["value"])

    return workbook_config


def _validate_runtime_structure(model_params: Mapping[str, Any]) -> dict[str, Any]:
    workbook_config = _validate_workbook_config(model_params)
    state_columns = list(workbook_config["state_columns"])
    composite_variables = list(workbook_config["composite_variables"])
    measured_output_columns = list(model_params["measured_output_columns"])
    process_names = [str(process["name"]) for process in workbook_config["processes"]]
    process_types = list(model_params["process_types"])
    operational_columns = list(model_params["operational_columns"])
    influent_state_ranges = dict(model_params["influent_state_ranges"])
    operational_ranges = dict(model_params["operational_ranges"])

    _validate_unique_names(measured_output_columns, "measured_output_columns")
    _validate_unique_names(process_names, "process names")
    _validate_unique_names(operational_columns, "operational_columns")

    if measured_output_columns != composite_variables:
        raise ValueError(
            "asm2d_tcn_simulation measured_output_columns must match workbook composite_variables for the "
            "composite-only output contract."
        )

    if len(process_types) != len(process_names):
        raise ValueError("asm2d_tcn_simulation process_types must align with the configured process list.")

    missing_state_ranges = [state_name for state_name in state_columns if state_name not in influent_state_ranges]
    if missing_state_ranges:
        missing_display = ", ".join(missing_state_ranges)
        raise KeyError(f"asm2d_tcn_simulation missing influent_state_ranges for: {missing_display}")

    missing_operational_ranges = [name for name in operational_columns if name not in operational_ranges]
    if missing_operational_ranges:
        missing_display = ", ".join(missing_operational_ranges)
        raise KeyError(f"asm2d_tcn_simulation missing operational_ranges for: {missing_display}")

    return {
        "workbook_config": workbook_config,
        "state_columns": state_columns,
        "measured_output_columns": measured_output_columns,
        "process_names": process_names,
        "process_types": process_types,
        "operational_columns": operational_columns,
        "influent_state_ranges": influent_state_ranges,
        "operational_ranges": operational_ranges,
    }


def _validate_unique_names(names: list[str], name_type: str) -> None:
    if not names:
        raise ValueError(f"{name_type} must not be empty.")

    duplicates = sorted({name for name in names if names.count(name) > 1})
    if duplicates:
        duplicate_display = ", ".join(duplicates)
        raise ValueError(f"asm2d_tcn_simulation {name_type} contains duplicates: {duplicate_display}")


def _build_parameter_reference_map(parameter_rows: list[Mapping[str, Any]]) -> dict[str, str]:
    value_column_letter = get_column_letter(PARAMETER_VALUE_COLUMN_INDEX)
    parameter_refs: dict[str, str] = {}

    for row_number, parameter_row in enumerate(parameter_rows, start=2):
        excel_name = str(parameter_row["excel_name"])
        parameter_refs[excel_name] = f"'{PARAMETER_SHEET_NAME}'!${value_column_letter}${row_number}"

    return parameter_refs


def _build_parameter_value_map(parameter_rows: list[Mapping[str, Any]]) -> dict[str, float]:
    return {str(parameter_row["excel_name"]): float(parameter_row["value"]) for parameter_row in parameter_rows}


def _evaluate_numeric_expression(expression: str | float | int, parameter_values: Mapping[str, float]) -> float:
    if isinstance(expression, (int, float)):
        return float(expression)

    formatted_expression = str(expression).format_map(parameter_values)
    return float(eval(formatted_expression, {"__builtins__": {}}, {}))


def _build_state_index(state_columns: list[str]) -> dict[str, int]:
    return {name: position for position, name in enumerate(state_columns)}


def _sample_named_ranges(
    rng: np.random.Generator,
    sample_count: int,
    ordered_names: list[str],
    ranges: Mapping[str, Any],
) -> np.ndarray:
    sampled = np.zeros((sample_count, len(ordered_names)), dtype=float)
    for column_index, column_name in enumerate(ordered_names):
        lower_bound, upper_bound = ranges[column_name]
        sampled[:, column_index] = rng.uniform(float(lower_bound), float(upper_bound), sample_count)

    return sampled


def _monod(numerator: float, half_saturation: float) -> float:
    return float(numerator) / max(float(numerator) + float(half_saturation), 1e-9)


def _ratio(numerator: float, denominator: float) -> float:
    return float(numerator) / max(float(denominator), 1e-9)


def _share(numerator: float, denominator_a: float, denominator_b: float) -> float:
    return float(numerator) / max(float(denominator_a) + float(denominator_b), 1e-9)


def _derive_tss_state(
    state: np.ndarray,
    state_index: Mapping[str, int],
    parameter_values: Mapping[str, float],
) -> float:
    return float(
        sum(
            state[state_index[state_name]] * _evaluate_numeric_expression(factor_expression, parameter_values)
            for state_name, factor_expression in TSS_CONTINUITY_TERMS.items()
        )
    )


def _build_influent_state_sample(
    sampled_state: np.ndarray,
    state_index: Mapping[str, int],
    parameter_values: Mapping[str, float],
) -> np.ndarray:
    influent_state = np.maximum(sampled_state.copy(), 0.0)
    influent_state[state_index["X_TSS"]] = _derive_tss_state(influent_state, state_index, parameter_values)
    return influent_state


def _build_initial_guess(
    influent_state: np.ndarray,
    aeration: float,
    state_index: Mapping[str, int],
    model_params: Mapping[str, Any],
    previous_solution: np.ndarray | None = None,
) -> np.ndarray:
    hyperparameters = model_params["hyperparameters"]
    aeration_model = model_params["aeration_model"]
    state_floor = float(hyperparameters["state_floor"])
    state_ceiling = float(hyperparameters["state_ceiling"])
    guess = np.clip(influent_state.copy(), state_floor, state_ceiling)

    if previous_solution is not None:
        guess = np.clip(0.55 * previous_solution + 0.45 * guess, state_floor, state_ceiling)

    guess[state_index["S_A"]] *= 0.7
    guess[state_index["S_F"]] *= 0.7
    guess[state_index["X_H"]] = max(guess[state_index["X_H"]], guess[state_index["X_S"]] * 0.25)
    guess[state_index["X_PAO"]] = max(guess[state_index["X_PAO"]], guess[state_index["X_PP"]] * 1.2)
    guess[state_index["X_AOB"]] = max(guess[state_index["X_AOB"]], guess[state_index["S_NH4"]] * 0.03)
    guess[state_index["X_NOB"]] = max(guess[state_index["X_NOB"]], guess[state_index["S_NO2"]] * 0.1)

    kla = float(aeration_model["kla_base"]) + float(aeration_model["kla_per_aeration"]) * max(float(aeration), 0.0)
    do_saturation = float(aeration_model["do_saturation"])
    guess[state_index["S_O2"]] = np.clip(
        (guess[state_index["S_O2"]] + kla * do_saturation) / max(1.0 + kla, 1e-9),
        state_floor,
        do_saturation,
    )

    return guess


def _compute_process_rates(
    state: np.ndarray,
    model_params: Mapping[str, Any],
    state_index: Mapping[str, int],
    parameter_values: Mapping[str, float],
) -> np.ndarray:
    s_a = state[state_index["S_A"]]
    s_f = state[state_index["S_F"]]
    s_nh4 = state[state_index["S_NH4"]]
    s_no2 = state[state_index["S_NO2"]]
    s_no3 = state[state_index["S_NO3"]]
    s_po4 = state[state_index["S_PO4"]]
    s_alk = state[state_index["S_ALK"]]
    s_o2 = state[state_index["S_O2"]]
    x_s = state[state_index["X_S"]]
    x_h = state[state_index["X_H"]]
    x_pao = state[state_index["X_PAO"]]
    x_pp = state[state_index["X_PP"]]
    x_pha = state[state_index["X_PHA"]]
    x_aob = state[state_index["X_AOB"]]
    x_nob = state[state_index["X_NOB"]]
    x_meoh = state[state_index["X_MeOH"]]
    x_mep = state[state_index["X_MeP"]]

    xs_to_xh = _ratio(x_s, x_h)
    hydrolysis_availability = xs_to_xh / max(parameter_values["K_X"] + xs_to_xh, 1e-9)
    oxygen_acceptor = _monod(s_o2, parameter_values["K_O2"])
    oxygen_limitation = parameter_values["K_O2"] / max(parameter_values["K_O2"] + s_o2, 1e-9)
    nox_total = s_no3 + s_no2
    nitrate_share = _share(s_no3, s_no3, s_no2)
    nitrite_share = _share(s_no2, s_no3, s_no2)
    alk_term = _monod(s_alk, parameter_values["K_ALK"])
    ammonium_term = _monod(s_nh4, parameter_values["K_NH4"])
    phosphate_term = _monod(s_po4, parameter_values["K_PO4"])
    pao_ratio_pp = _ratio(x_pp, x_pao)
    pao_ratio_pha = _ratio(x_pha, x_pao)
    pp_capacity = max(parameter_values["K_max"] - pao_ratio_pp, 0.0)

    rho5 = (
        parameter_values["mu_H"]
        * oxygen_acceptor
        * _monod(s_f, parameter_values["K_F"])
        * ammonium_term
        * phosphate_term
        * alk_term
        * x_h
    )
    rho6 = (
        parameter_values["mu_H"]
        * oxygen_acceptor
        * _monod(s_a, parameter_values["K_A"])
        * ammonium_term
        * phosphate_term
        * alk_term
        * x_h
    )
    rho14 = (
        parameter_values["q_PP"]
        * oxygen_acceptor
        * phosphate_term
        * alk_term
        * (_ratio(x_pha, x_pao) / max(parameter_values["K_PHA"] + _ratio(x_pha, x_pao), 1e-9))
        * (pp_capacity / max(parameter_values["K_IPP"] + pp_capacity, 1e-9))
        * x_pao
    )
    rho17 = (
        parameter_values["mu_PAO"]
        * oxygen_acceptor
        * ammonium_term
        * phosphate_term
        * alk_term
        * (_ratio(x_pha, x_pao) / max(parameter_values["K_PHA"] + _ratio(x_pha, x_pao), 1e-9))
        * x_pao
    )

    process_rates = np.array(
        [
            parameter_values["K_h"] * oxygen_acceptor * hydrolysis_availability * x_h,
            parameter_values["K_h"]
            * parameter_values["eta_NO2"]
            * oxygen_limitation
            * _monod(s_no2, parameter_values["K_NO2"])
            * nitrite_share
            * hydrolysis_availability
            * x_h,
            parameter_values["K_h"]
            * parameter_values["eta_NO3"]
            * oxygen_limitation
            * _monod(s_no3, parameter_values["K_NO3"])
            * nitrate_share
            * hydrolysis_availability
            * x_h,
            parameter_values["K_h"]
            * parameter_values["eta_fe"]
            * oxygen_limitation
            * (parameter_values["K_NOX"] / max(parameter_values["K_NOX"] + nox_total, 1e-9))
            * hydrolysis_availability
            * x_h,
            rho5,
            rho6,
            rho5
            * parameter_values["eta_NO3"]
            * oxygen_limitation
            * _monod(s_no3, parameter_values["K_NO3"])
            * nitrate_share,
            rho5
            * parameter_values["eta_NO2"]
            * oxygen_limitation
            * _monod(s_no2, parameter_values["K_NO2"])
            * nitrite_share,
            rho6
            * parameter_values["eta_NO3"]
            * oxygen_limitation
            * _monod(s_no3, parameter_values["K_NO3"])
            * nitrate_share,
            rho6
            * parameter_values["eta_NO2"]
            * oxygen_limitation
            * _monod(s_no2, parameter_values["K_NO2"])
            * nitrite_share,
            parameter_values["q_fe"]
            * parameter_values["mu_H"]
            * oxygen_limitation
            * (parameter_values["K_NOX"] / max(parameter_values["K_NOX"] + nox_total, 1e-9))
            * _monod(s_f, parameter_values["K_fe"])
            * alk_term
            * x_h,
            parameter_values["b_H"] * x_h,
            parameter_values["q_PHA"]
            * _monod(s_a, parameter_values["K_A"])
            * alk_term
            * (pao_ratio_pp / max(parameter_values["K_PP"] + pao_ratio_pp, 1e-9))
            * x_pao,
            rho14,
            rho14
            * parameter_values["eta_NO3"]
            * oxygen_limitation
            * _monod(s_no3, parameter_values["K_NO3"])
            * nitrate_share,
            rho14
            * parameter_values["eta_NO2"]
            * oxygen_limitation
            * _monod(s_no2, parameter_values["K_NO2"])
            * nitrite_share,
            rho17,
            rho17
            * parameter_values["eta_NO3"]
            * oxygen_limitation
            * _monod(s_no3, parameter_values["K_NO3"])
            * nitrate_share,
            rho17
            * parameter_values["eta_NO2"]
            * oxygen_limitation
            * _monod(s_no2, parameter_values["K_NO2"])
            * nitrite_share,
            parameter_values["b_PAO"] * x_pao * alk_term,
            parameter_values["b_PP"] * x_pp * alk_term,
            parameter_values["b_PHA"] * x_pha * alk_term,
            parameter_values["mu_AOB"]
            * _monod(s_o2, parameter_values["K_O2_AOB"])
            * _monod(s_nh4, parameter_values["K_NH4_AOB"])
            * phosphate_term
            * alk_term
            * x_aob,
            parameter_values["mu_NOB"]
            * _monod(s_o2, parameter_values["K_O2_NOB"])
            * _monod(s_no2, parameter_values["K_NO2_NOB"])
            * phosphate_term
            * alk_term
            * x_nob,
            parameter_values["b_AOB"] * x_aob,
            parameter_values["b_NOB"] * x_nob,
            parameter_values["k_PRE"] * s_po4 * x_meoh,
            parameter_values["k_RED"] * x_mep * alk_term,
        ],
        dtype=float,
    )
    return np.clip(process_rates, 0.0, float(model_params["hyperparameters"]["max_process_rate"]))


def _compute_aeration_flux(state: np.ndarray, aeration: float, state_index: Mapping[str, int], model_params: Mapping[str, Any]) -> float:
    aeration_model = model_params["aeration_model"]
    kla = float(aeration_model["kla_base"]) + float(aeration_model["kla_per_aeration"]) * max(float(aeration), 0.0)
    do_saturation = float(aeration_model["do_saturation"])
    return kla * (do_saturation - state[state_index["S_O2"]])


def _steady_state_residuals(
    state: np.ndarray,
    influent_state: np.ndarray,
    hrt_hours: float,
    aeration: float,
    matrix_bundle: Mapping[str, Any],
    model_params: Mapping[str, Any],
    parameter_values: Mapping[str, float],
) -> np.ndarray:
    dilution_rate = 24.0 / max(float(hrt_hours), 1e-6)
    state_index = dict(matrix_bundle["state_index"])
    residual = dilution_rate * (influent_state - state)
    process_rates = _compute_process_rates(state, model_params, state_index, parameter_values)
    residual += process_rates @ np.asarray(matrix_bundle["petersen_matrix"], dtype=float)
    residual[state_index["S_O2"]] += _compute_aeration_flux(state, aeration, state_index, model_params)
    return residual


def simulate_asm2d_tcn_steady_state(
    *,
    influent_state: np.ndarray,
    hrt_hours: float,
    aeration: float,
    model_params: Mapping[str, Any],
    matrix_bundle: Mapping[str, Any] | None = None,
    previous_solution: np.ndarray | None = None,
) -> tuple[np.ndarray, dict[str, float | bool | int]]:
    """Approximate the ASM2d-TCN steady state by relaxed fixed-point iteration."""

    matrix_bundle = matrix_bundle if matrix_bundle is not None else get_asm2d_tcn_matrices(model_params)
    runtime = _validate_runtime_structure(model_params)
    parameter_values = _build_parameter_value_map(runtime["workbook_config"]["parameters"])
    state_index = dict(matrix_bundle["state_index"])
    hyperparameters = model_params["hyperparameters"]
    base_iterations = int(hyperparameters["fixed_point_iterations"])
    base_relaxation = float(hyperparameters["fixed_point_relaxation"])
    state_floor = float(hyperparameters["state_floor"])
    state_ceiling = float(hyperparameters["state_ceiling"])

    best_state = None
    best_residual = None
    best_iteration_count = 0
    for iteration_count, relaxation in ((base_iterations, base_relaxation), (base_iterations * 2, base_relaxation * 0.5)):
        state = _build_initial_guess(
            influent_state,
            aeration,
            state_index,
            model_params,
            previous_solution=previous_solution,
        )
        for _ in range(iteration_count):
            residual = _steady_state_residuals(
                state,
                influent_state,
                hrt_hours,
                aeration,
                matrix_bundle,
                model_params,
                parameter_values,
            )
            dilution_rate = 24.0 / max(float(hrt_hours), 1e-6)
            state = np.clip(
                state + relaxation * (residual / max(dilution_rate, 1e-9)),
                state_floor,
                state_ceiling,
            )
            if not np.all(np.isfinite(state)):
                raise RuntimeError("asm2d_tcn_simulation produced non-finite state values during fixed-point iteration.")

        final_residual = _steady_state_residuals(
            state,
            influent_state,
            hrt_hours,
            aeration,
            matrix_bundle,
            model_params,
            parameter_values,
        )
        if best_residual is None or np.max(np.abs(final_residual)) < np.max(np.abs(best_residual)):
            best_state = state
            best_residual = final_residual
            best_iteration_count = iteration_count

    assert best_state is not None
    assert best_residual is not None
    diagnostics: dict[str, float | bool | int] = {
        "success": bool(np.max(np.abs(best_residual)) <= float(hyperparameters["acceptance_residual_max"])),
        "status": 1,
        "iterations": best_iteration_count,
        "residual_l2": float(np.linalg.norm(best_residual)),
        "residual_max": float(np.max(np.abs(best_residual))),
    }
    return best_state, diagnostics


def _compute_measured_output_values(state: np.ndarray, matrix_bundle: Mapping[str, Any]) -> np.ndarray:
    return np.asarray(matrix_bundle["composition_matrix"], dtype=float) @ state


def _generate_asm2d_tcn_dataset_chunk(
    *,
    chunk_start: int,
    chunk_size: int,
    chunk_seed: int,
    model_params: Mapping[str, Any],
    matrix_bundle: Mapping[str, Any],
    runtime: Mapping[str, Any],
) -> tuple[int, np.ndarray, np.ndarray, np.ndarray]:
    configured_hyperparameters = model_params["hyperparameters"]
    max_sample_attempts = int(configured_hyperparameters["max_sample_attempts"])
    state_columns = list(runtime["state_columns"])
    operational_columns = list(runtime["operational_columns"])
    measured_output_columns = list(runtime["measured_output_columns"])
    state_index = dict(matrix_bundle["state_index"])
    parameter_values = _build_parameter_value_map(runtime["workbook_config"]["parameters"])
    rng = np.random.default_rng(chunk_seed)

    influent_states = np.zeros((chunk_size, len(state_columns)), dtype=float)
    operational = np.zeros((chunk_size, len(operational_columns)), dtype=float)
    measured_outputs = np.zeros((chunk_size, len(measured_output_columns)), dtype=float)

    previous_solution: np.ndarray | None = None
    for local_index in range(chunk_size):
        last_error: RuntimeError | None = None
        for _attempt_index in range(max_sample_attempts):
            sampled_state = _sample_named_ranges(rng, 1, state_columns, runtime["influent_state_ranges"])[0]
            candidate_influent = _build_influent_state_sample(sampled_state, state_index, parameter_values)
            candidate_operational = _sample_named_ranges(rng, 1, operational_columns, runtime["operational_ranges"])[0]

            try:
                effluent_state, _ = simulate_asm2d_tcn_steady_state(
                    influent_state=candidate_influent,
                    hrt_hours=float(candidate_operational[0]),
                    aeration=float(candidate_operational[1]),
                    model_params=model_params,
                    matrix_bundle=matrix_bundle,
                    previous_solution=previous_solution,
                )
            except RuntimeError as error:
                last_error = error
                continue

            if not np.all(np.isfinite(effluent_state)):
                last_error = RuntimeError("asm2d_tcn_simulation produced non-finite effluent states.")
                continue

            previous_solution = effluent_state
            influent_states[local_index] = candidate_influent
            operational[local_index] = candidate_operational
            measured_outputs[local_index] = _compute_measured_output_values(effluent_state, matrix_bundle)
            break
        else:
            raise RuntimeError(
                "asm2d_tcn_simulation failed to generate a valid sample after "
                f"{max_sample_attempts} attempts at sample index {chunk_start + local_index}."
            ) from last_error

    return chunk_start, influent_states, operational, measured_outputs


def _resolve_parallel_workers(requested_workers: int, sample_count: int) -> int:
    if requested_workers < 0:
        raise ValueError("parallel_workers must be greater than or equal to 0.")

    if sample_count <= 1:
        return 1

    available_workers = os.cpu_count() or 1
    if requested_workers == 0:
        requested_workers = max(available_workers - 1, 1)

    return min(max(requested_workers, 1), available_workers, sample_count)


def _resolve_parallel_chunk_size(requested_chunk_size: int) -> int:
    if requested_chunk_size < 1:
        raise ValueError("parallel_chunk_size must be at least 1.")

    return requested_chunk_size


def _write_parameter_sheet(worksheet, parameter_rows: list[Mapping[str, Any]]) -> None:
    worksheet.freeze_panes = "A2"
    headers = ["category", "symbol", "excel_name", "description", "value", "unit"]
    _write_header_row(worksheet, headers)

    previous_category = None
    for row_number, parameter_row in enumerate(parameter_rows, start=2):
        current_category = str(parameter_row["category"])
        row_values = [
            current_category,
            str(parameter_row["symbol"]),
            str(parameter_row["excel_name"]),
            str(parameter_row["description"]),
            float(parameter_row["value"]),
            str(parameter_row["unit"]),
        ]

        for column_number, value in enumerate(row_values, start=1):
            cell = worksheet.cell(row=row_number, column=column_number, value=value)
            if column_number == PARAMETER_VALUE_COLUMN_INDEX:
                cell.number_format = "0.###############"

        if current_category != previous_category:
            for column_number in range(1, len(headers) + 1):
                worksheet.cell(row=row_number, column=column_number).fill = SECTION_FILL
        previous_category = current_category


def _write_stoichiometric_sheet(
    worksheet,
    workbook_config: Mapping[str, Any],
    parameter_refs: Mapping[str, str],
) -> None:
    worksheet.freeze_panes = "C2"
    state_columns = list(workbook_config["state_columns"])
    processes = list(workbook_config["processes"])
    headers = ["process_index", "process"] + state_columns
    _write_header_row(worksheet, headers)
    state_column_index = {state_name: position for position, state_name in enumerate(state_columns, start=3)}

    for row_number, process in enumerate(processes, start=2):
        worksheet.cell(row=row_number, column=1, value=int(process["index"]))
        worksheet.cell(row=row_number, column=2, value=str(process["name"]))
        direct_coefficients = STOICHIOMETRIC_COEFFICIENTS[row_number - 2]["coefficients"]

        for state_name in state_columns:
            column_number = state_column_index[state_name]
            cell = worksheet.cell(row=row_number, column=column_number)

            if state_name in direct_coefficients:
                cell.value = _format_formula(direct_coefficients[state_name], parameter_refs)
                continue

            if state_name == "S_NH4":
                cell.value = _build_weighted_formula(
                    row_number,
                    state_column_index,
                    NITROGEN_CONTINUITY_TERMS,
                    parameter_refs,
                    negate=True,
                )
                continue

            if state_name == "S_PO4":
                cell.value = _build_weighted_formula(
                    row_number,
                    state_column_index,
                    PHOSPHORUS_CONTINUITY_TERMS,
                    parameter_refs,
                    negate=True,
                )
                continue

            if state_name == "S_ALK":
                cell.value = _build_alkalinity_formula(row_number, state_column_index)
                continue

            if state_name == "X_TSS":
                cell.value = _build_weighted_formula(
                    row_number,
                    state_column_index,
                    TSS_CONTINUITY_TERMS,
                    parameter_refs,
                    negate=False,
                )


def _write_composition_sheet(
    worksheet,
    workbook_config: Mapping[str, Any],
    parameter_refs: Mapping[str, str],
) -> None:
    worksheet.freeze_panes = "D2"
    dissolved_state_columns = list(workbook_config["dissolved_state_columns"])
    particulate_state_columns = list(workbook_config["particulate_state_columns"])
    state_columns = list(workbook_config["state_columns"])
    composite_variables = list(workbook_config["composite_variables"])
    state_units = dict(workbook_config["state_units"])
    headers = ["state_group", "state_variable", "unit"] + composite_variables
    _write_header_row(worksheet, headers)
    composite_column_index = {name: position for position, name in enumerate(composite_variables, start=4)}

    for row_number, state_name in enumerate(state_columns, start=2):
        state_group = "Dissolved" if state_name in dissolved_state_columns else "Particulate"
        worksheet.cell(row=row_number, column=1, value=state_group)
        worksheet.cell(row=row_number, column=2, value=state_name)
        worksheet.cell(row=row_number, column=3, value=state_units[state_name])

        for composite_name, expression in COMPOSITION_FORMULAS.get(state_name, {}).items():
            worksheet.cell(
                row=row_number,
                column=composite_column_index[composite_name],
                value=_format_formula(expression, parameter_refs),
            )


def _write_header_row(worksheet, headers: list[str]) -> None:
    for column_number, header in enumerate(headers, start=1):
        cell = worksheet.cell(row=1, column=column_number, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _format_formula(expression: str | float | int, parameter_refs: Mapping[str, str]) -> str:
    if isinstance(expression, (int, float)):
        return f"={expression:g}"

    formatted_expression = str(expression).format_map(parameter_refs)
    if formatted_expression.startswith("="):
        return formatted_expression

    return f"={formatted_expression}"


def _build_weighted_formula(
    row_number: int,
    state_column_index: Mapping[str, int],
    factor_terms: Mapping[str, str],
    parameter_refs: Mapping[str, str],
    *,
    negate: bool,
) -> str:
    terms: list[str] = []
    for state_name, factor_expression in factor_terms.items():
        cell_reference = f"{get_column_letter(state_column_index[state_name])}{row_number}"
        formatted_factor = factor_expression.format_map(parameter_refs)
        if formatted_factor == "1":
            terms.append(cell_reference)
        else:
            terms.append(f"{cell_reference}*({formatted_factor})")

    if not terms:
        return "=0"

    expression = "+".join(terms)
    if negate:
        return f"=-({expression})"

    return f"={expression}"


def _build_alkalinity_formula(row_number: int, state_column_index: Mapping[str, int]) -> str:
    ammonium_ref = f"{get_column_letter(state_column_index['S_NH4'])}{row_number}"
    nitrite_ref = f"{get_column_letter(state_column_index['S_NO2'])}{row_number}"
    nitrate_ref = f"{get_column_letter(state_column_index['S_NO3'])}{row_number}"
    phosphate_ref = f"{get_column_letter(state_column_index['S_PO4'])}{row_number}"
    return f"={ammonium_ref}/14-{nitrite_ref}/14-{nitrate_ref}/14+{phosphate_ref}/31"


def _auto_size_columns(worksheet) -> None:
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is None:
                continue
            max_length = max(max_length, len(str(cell.value)))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 48)


__all__ = [
    "build_asm2d_tcn_workbook",
    "build_asm2d_tcn_metadata",
    "create_asm2d_tcn_workbook",
    "generate_asm2d_tcn_dataset",
    "get_asm2d_tcn_matrices",
    "load_asm2d_tcn_simulation_params",
    "resolve_asm2d_tcn_simulation_artifact_paths",
    "resolve_asm2d_tcn_workbook_path",
    "run_asm2d_tcn_simulation",
    "simulate_asm2d_tcn_steady_state",
]