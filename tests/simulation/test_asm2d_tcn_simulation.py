"""Workbook and runtime contract tests for the ASM2d-TCN reference model."""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import numpy as np
from openpyxl import load_workbook

from src.models.simulation.asm2d_tcn_simulation import (
    _build_influent_state_sample,
    _build_parameter_value_map,
    generate_asm2d_tcn_dataset,
    get_asm2d_tcn_matrices,
    create_asm2d_tcn_workbook,
    load_asm2d_tcn_simulation_params,
    resolve_asm2d_tcn_simulation_artifact_paths,
    resolve_asm2d_tcn_workbook_path,
    run_asm2d_tcn_simulation,
    simulate_asm2d_tcn_steady_state,
    sweep_asm2d_tcn_operating_space,
)


def _column_index_by_header(worksheet) -> dict[str, int]:
    return {
        str(cell.value): index
        for index, cell in enumerate(worksheet[1], start=1)
        if cell.value is not None
    }


def _row_index_by_value(worksheet, column_number: int) -> dict[str, int]:
    index: dict[str, int] = {}
    for row_number in range(2, worksheet.max_row + 1):
        value = worksheet.cell(row=row_number, column=column_number).value
        if value is not None:
            index[str(value)] = row_number
    return index


def _build_midpoint_influent_state(model_params: dict[str, object]) -> np.ndarray:
    state_columns = list(model_params["workbook"]["state_columns"])
    state_index = {name: position for position, name in enumerate(state_columns)}
    parameter_values = _build_parameter_value_map(model_params["workbook"]["parameters"])
    midpoint_sample = np.array(
        [
            np.mean(model_params["influent_state_ranges"][state_name])
            for state_name in state_columns
        ],
        dtype=float,
    )
    return _build_influent_state_sample(midpoint_sample, state_index, parameter_values)


class Asm2dTcnWorkbookTests(unittest.TestCase):
    def test_resolve_workbook_path_uses_configured_location(self) -> None:
        workbook_path = resolve_asm2d_tcn_workbook_path()

        self.assertTrue(workbook_path.as_posix().endswith("data/asm2d-tcn/asm2d_tcn_workbook.xlsx"))

    def test_workbook_config_contains_expected_dimensions(self) -> None:
        params = load_asm2d_tcn_simulation_params()
        workbook_config = params["workbook"]

        self.assertEqual(
            workbook_config["sheets"],
            ["stoichiometric_matrix", "composition_matrix", "parameter_table"],
        )
        self.assertEqual(len(workbook_config["processes"]), 28)
        self.assertEqual(len(workbook_config["state_columns"]), 21)
        self.assertEqual(len(workbook_config["composite_variables"]), 6)

    def test_create_workbook_writes_required_sheets(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = create_asm2d_tcn_workbook(Path(tmp_dir) / "asm2d_tcn.xlsx")
            workbook = load_workbook(workbook_path, data_only=False)

        self.assertEqual(
            workbook.sheetnames,
            ["stoichiometric_matrix", "composition_matrix", "parameter_table"],
        )

    def test_stoichiometric_matrix_contains_formula_links(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = create_asm2d_tcn_workbook(Path(tmp_dir) / "asm2d_tcn.xlsx")
            workbook = load_workbook(workbook_path, data_only=False)

        worksheet = workbook["stoichiometric_matrix"]
        header_index = _column_index_by_header(worksheet)
        process_row_index = _row_index_by_value(worksheet, 2)

        aerobic_hydrolysis_row = process_row_index["Aerobic hydrolysis"]
        precipitation_row = process_row_index["Precipitation"]
        aob_growth_row = process_row_index["Aerobic growth of X_AOB"]

        self.assertIn("parameter_table", str(worksheet.cell(aerobic_hydrolysis_row, header_index["S_F"]).value))
        self.assertIn("parameter_table", str(worksheet.cell(aerobic_hydrolysis_row, header_index["S_NH4"]).value))
        self.assertIn("parameter_table", str(worksheet.cell(aob_growth_row, header_index["S_NO2"]).value))
        self.assertEqual(worksheet.cell(precipitation_row, header_index["S_PO4"]).value, "=-1")
        self.assertEqual(worksheet.cell(precipitation_row, header_index["X_TSS"]).value, "=1.42")

    def test_composition_matrix_contains_formula_links(self) -> None:
        with tempfile.TemporaryDirectory() as tmp_dir:
            workbook_path = create_asm2d_tcn_workbook(Path(tmp_dir) / "asm2d_tcn.xlsx")
            workbook = load_workbook(workbook_path, data_only=False)

        worksheet = workbook["composition_matrix"]
        header_index = _column_index_by_header(worksheet)
        state_row_index = _row_index_by_value(worksheet, 2)

        self.assertIn("parameter_table", str(worksheet.cell(state_row_index["X_H"], header_index["TN"]).value))
        self.assertIn("parameter_table", str(worksheet.cell(state_row_index["X_MeP"], header_index["TP"]).value))
        self.assertEqual(worksheet.cell(state_row_index["X_TSS"], header_index["TSS"]).value, "=1")


class Asm2dTcnSimulationTests(unittest.TestCase):
    def test_resolve_simulation_artifact_paths_use_requested_folder(self) -> None:
        dataset_path, metadata_path, dataset_relative = resolve_asm2d_tcn_simulation_artifact_paths(
            timestamp="20260330_000000"
        )

        self.assertTrue(dataset_path.as_posix().endswith("data/asm2d-tcn/simulation/data_20260330_000000.csv"))
        self.assertTrue(metadata_path.as_posix().endswith("data/asm2d-tcn/simulation/metadata_20260330_000000.json"))
        self.assertEqual(dataset_relative, "data/asm2d-tcn/simulation/data_20260330_000000.csv")

    def test_numeric_matrices_have_expected_shapes(self) -> None:
        params = load_asm2d_tcn_simulation_params()
        matrix_bundle = get_asm2d_tcn_matrices(params)

        self.assertEqual(matrix_bundle["petersen_matrix"].shape, (28, 21))
        self.assertEqual(matrix_bundle["composition_matrix"].shape, (6, 21))
        self.assertEqual(matrix_bundle["measured_output_columns"], params["measured_output_columns"])

    def test_generate_dataset_reports_composites_only(self) -> None:
        params = load_asm2d_tcn_simulation_params()
        dataset, metadata, matrix_bundle = generate_asm2d_tcn_dataset(
            model_params=params,
            n_samples=12,
            random_seed=7,
            parallel_workers=1,
        )
        expected_independent = metadata["independent_columns"]
        expected_dependent = metadata["dependent_columns"]

        self.assertEqual(dataset.shape, (12, len(expected_independent) + len(expected_dependent)))
        self.assertEqual(list(dataset.columns), expected_independent + expected_dependent)
        self.assertEqual(expected_dependent, [f"Out_{name}" for name in params["measured_output_columns"]])
        self.assertFalse(any(column_name.startswith("Out_S_") or column_name.startswith("Out_X_") for column_name in expected_dependent))
        self.assertEqual(matrix_bundle["petersen_matrix"].shape, (28, 21))
        self.assertEqual(matrix_bundle["composition_matrix"].shape, (6, 21))

    def test_single_operating_point_solves_to_small_residual(self) -> None:
        params = load_asm2d_tcn_simulation_params()
        influent_state = _build_midpoint_influent_state(params)

        solution, diagnostics = simulate_asm2d_tcn_steady_state(
            influent_state=influent_state,
            hrt_hours=24.0,
            aeration=1.5,
            model_params=params,
        )

        self.assertTrue(diagnostics["success"])
        self.assertLess(diagnostics["residual_max"], 1e-5)
        self.assertTrue((solution >= 0.0).all())

    def test_steady_state_responds_to_aeration_and_hrt(self) -> None:
        params = load_asm2d_tcn_simulation_params()
        matrix_bundle = get_asm2d_tcn_matrices(params)
        influent_state = _build_midpoint_influent_state(params)
        state_index = dict(matrix_bundle["state_index"])
        output_index = {
            name: position for position, name in enumerate(matrix_bundle["measured_output_columns"])
        }

        low_aeration_state, _ = simulate_asm2d_tcn_steady_state(
            influent_state=influent_state,
            hrt_hours=24.0,
            aeration=0.75,
            model_params=params,
            matrix_bundle=matrix_bundle,
        )
        high_aeration_state, _ = simulate_asm2d_tcn_steady_state(
            influent_state=influent_state,
            hrt_hours=24.0,
            aeration=2.25,
            model_params=params,
            matrix_bundle=matrix_bundle,
        )

        self.assertGreater(high_aeration_state[state_index["S_O2"]], low_aeration_state[state_index["S_O2"]])
        self.assertLess(high_aeration_state[state_index["S_NH4"]], low_aeration_state[state_index["S_NH4"]])

        low_hrt_state, _ = simulate_asm2d_tcn_steady_state(
            influent_state=influent_state,
            hrt_hours=12.0,
            aeration=1.5,
            model_params=params,
            matrix_bundle=matrix_bundle,
        )
        high_hrt_state, _ = simulate_asm2d_tcn_steady_state(
            influent_state=influent_state,
            hrt_hours=36.0,
            aeration=1.5,
            model_params=params,
            matrix_bundle=matrix_bundle,
        )
        low_hrt_outputs = matrix_bundle["composition_matrix"] @ low_hrt_state
        high_hrt_outputs = matrix_bundle["composition_matrix"] @ high_hrt_state

        self.assertLess(high_hrt_outputs[output_index["COD"]], low_hrt_outputs[output_index["COD"]])
        self.assertLess(high_hrt_state[state_index["S_NH4"]], low_hrt_state[state_index["S_NH4"]])

    def test_generate_dataset_retries_solver_failures(self) -> None:
        params = load_asm2d_tcn_simulation_params()
        original_solver = simulate_asm2d_tcn_steady_state
        fail_once = {"remaining": 1}

        def flaky_solver(*args, **kwargs):
            if fail_once["remaining"] > 0:
                fail_once["remaining"] -= 1
                raise RuntimeError("transient steady-state failure")
            return original_solver(*args, **kwargs)

        with patch(
            "src.models.simulation.asm2d_tcn_simulation.simulate_asm2d_tcn_steady_state",
            side_effect=flaky_solver,
        ):
            dataset, metadata, _ = generate_asm2d_tcn_dataset(
                model_params=params,
                n_samples=1,
                random_seed=7,
                parallel_workers=1,
            )

        self.assertEqual(dataset.shape[0], 1)
        self.assertEqual(metadata["n_samples"], 1)
        self.assertEqual(fail_once["remaining"], 0)

    def test_run_simulation_returns_notebook_facing_bundle(self) -> None:
        result = run_asm2d_tcn_simulation(
            save_artifacts=False,
            n_samples=8,
            random_seed=5,
            parallel_workers=1,
        )

        self.assertIn("dataset", result)
        self.assertIn("metadata", result)
        self.assertIn("petersen_matrix", result)
        self.assertIn("composition_matrix", result)
        self.assertIn("matrix_bundle", result)
        self.assertIn("artifact_paths", result)
        self.assertEqual(result["artifact_paths"]["dataset_csv"], None)
        self.assertEqual(result["artifact_paths"]["metadata_json"], None)
        self.assertEqual(result["metadata"]["measured_output_columns"], ["COD", "TN", "TKN", "TP", "TSS", "VSS"])

    def test_run_simulation_can_return_in_memory_debug_payloads(self) -> None:
        params = load_asm2d_tcn_simulation_params()
        result = run_asm2d_tcn_simulation(
            save_artifacts=False,
            n_samples=4,
            random_seed=5,
            parallel_workers=1,
            include_debug_data=True,
            show_progress=False,
        )

        self.assertEqual(result["dataset"].shape[0], 4)
        self.assertIsNotNone(result["effluent_states"])
        self.assertIsNotNone(result["solver_diagnostics"])
        self.assertIsNotNone(result["solver_summary"])
        self.assertEqual(result["effluent_states"].shape, (4, len(params["workbook"]["state_columns"])))
        self.assertEqual(len(result["solver_diagnostics"]), 4)
        self.assertEqual(result["solver_summary"]["sample_count"], 4)
        self.assertIn("selected_strategy", result["solver_diagnostics"].columns)
        self.assertIn("dynamic_relaxation_used", result["solver_diagnostics"].columns)

    def test_operating_space_sweep_returns_calibration_summary(self) -> None:
        sweep_result = sweep_asm2d_tcn_operating_space(
            n_samples=16,
            random_seed=7,
            show_progress=False,
        )

        self.assertEqual(sweep_result["influent_states"].shape[0], 16)
        self.assertEqual(sweep_result["operating_conditions"].shape[0], 16)
        self.assertEqual(sweep_result["effluent_states"].shape[0], 16)
        self.assertEqual(len(sweep_result["solver_diagnostics"]), 16)
        self.assertEqual(sweep_result["summary"]["sample_count"], 16)
        self.assertGreaterEqual(sweep_result["summary"]["accepted_count"], 1)
        self.assertIn("residual_max_quantiles", sweep_result["summary"])
        self.assertIn("selected_strategy_counts", sweep_result["summary"])


if __name__ == "__main__":
    unittest.main()